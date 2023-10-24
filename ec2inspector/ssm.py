from .security import login_required, ssm_enabled
from flask import request, jsonify, session, abort, send_file, render_template
import uuid
import time
import datetime
import json
import hashlib
import boto3
import random
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO, StringIO
from botocore.session import Session as BotocoreSession
from functools import wraps
from .database.models import History, db
import jwt

SECRET_KEY = "+L2GKb/r96s^Yl<&j1zJ&CER+4\;9+@s"

def init_ssm(app):

    def get_boto3_session():
        botocore_session = BotocoreSession()
        botocore_session.set_config_variable('credentials_file', session.get('credentials_file'))
        botocore_session.set_config_variable('config_file', session.get('config_file'))
        boto3_session = boto3.Session(botocore_session=botocore_session, profile_name=session.get('selectedprofile'))

        return boto3_session
    
    # Send command to EC2 instance
    def send_command(instanceid, command, region, jsonifyno = "Yes"):
        while True:
            try:
                boto3_session = get_boto3_session()
                ssm_client = boto3_session.client('ssm', region_name=region)
                break
            except KeyError:
                time.sleep(1)

        try:    
            response = ssm_client.send_command(InstanceIds=[instanceid], DocumentName="AWS-RunShellScript", Parameters={"commands":[command]})
        except ssm_client.exceptions.InvalidInstanceId:
            return "Instance ID is not valid"

        command_id = response['Command']['CommandId']

        while True:
            try:
                output = ssm_client.get_command_invocation(
                    CommandId=command_id,
                    InstanceId=instanceid,
                )
                if output['Status'] == 'InProgress':
                    time.sleep(0.1)
                else:
                    break
            except ssm_client.exceptions.InvocationDoesNotExist:
                time.sleep(0.1)
            except ssm_client.exceptions.InvalidInstanceId:
                return "Instance ID is not valid"

        if jsonifyno == "no":
            return output['StandardOutputContent']
        else:
            return jsonify(output['StandardOutputContent'])

    
    # Send loop command EC2 instance
    def loop_command(file_kb, instanceid, file, region, jsonifyno = "Yes"):
        output = ""
        
        for i in range(0, int(file_kb), 23):
            commandloop = f'dd if=/tmp/{file} bs=1K skip={i} count=23'
            data = send_command(instanceid, commandloop, region, jsonifyno = "no")
            output += data

        command = f'rm -rf /tmp/{file}'
        send_command(instanceid, command, region)

        if jsonifyno == "no":
            return output
        else:
            return jsonify(output)
    
    @app.route('/get_data_ssm_check_permises', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_check_permises():
        return "True"

    @app.route('/get_data_ssm_testec2', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_testec2():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "echo 'SSM is working!'"

        '''
        Errores posibles:
        
        Tiene rol IAM?
        Tiene permisos?
        Tiene agente?
        Hay conectividad?
        '''

        return send_command(instanceid, command, region)
    
    @app.route('/get_data_ssm_install_dependences', methods=['HEAD'])
    @login_required
    @ssm_enabled
    def get_data_ssm_install_dependences():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "bash -c 'if command -v apt &> /dev/null; then dpkg -l sysstat net-tools &> /dev/null || (apt update && apt install sysstat net-tools -y); elif command -v yum &> /dev/null; then yum list installed sysstat net-tools &> /dev/null || yum install sysstat net-tools -y; elif command -v dnf &> /dev/null; then dnf list installed sysstat net-tools &> /dev/null || dnf install sysstat net-tools -y; elif command -v pacman &> /dev/null; then pacman -Q sysstat net-tools &> /dev/null || (pacman -Sy && pacman -S sysstat net-tools --noconfirm); elif command -v zypper &> /dev/null; then zypper se --installed-only sysstat net-tools &> /dev/null || (zypper refresh && zypper install sysstat net-tools -y); fi'"

        return send_command(instanceid, command, region)
    
    @app.route('/get_data_ssm_graphic', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_graphic():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "if [ $(cat /etc/os-release | sed -n 's/^PRETTY_NAME=\"\(.*\)\"/\\1/p') != '' ]; then cat /etc/os-release | sed -n 's/^PRETTY_NAME=\"\(.*\)\"/\\1/p'; else cat /etc/os-release | sed -n 's/^NAME=\"\(.*\)\"/\\1/p'; fi | tr -d '\n' && echo -n ';' && uptime -s | tr -d '\n' && echo -n ', ' && uptime -p | tr -d '\n' && echo -n ';' && top -bn1 | awk '/Cpu/ { print $2}' | tr -d '\n' && echo -n ';' && free | grep Mem | awk '{print $3/$2 * 100.0}' | tr -d '\\n'"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_disk', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_disk():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "u=\"$(cat /proc/sys/kernel/random/uuid)\";f=\"/tmp/diskstats_start_$u\";awk '{print$3,$6,$10}' /proc/diskstats>\"$f\";sleep 1;while read -r l;do d=$(echo \"$l\"|awk '{print$3}');rs=$(echo \"$l\"|awk '{print$6}');ws=$(echo \"$l\"|awk '{print$10}');if [ \"$d\" == \"\" ]||[ \"$rs\" == \"\" ]||[ \"$ws\" == \"\" ];then continue;fi;rss=$(awk -v d=\"$d\" '$1==d {print $2}' \"$f\");wss=$(awk -v d=\"$d\" '$1==d {print $3}' \"$f\");if [ -n \"$rss\" ]&&[ -n \"$wss\" ];then rk=$(((rs-rss)/2));wk=$(((ws-wss)/2));echo \"$l\"|awk -v rk=\"$rk\" -v wk=\"$wk\" '{print $3\",\"rk\",\"wk\",\"$4\",\"$8\",\"$7\",\"$11\",\"$6\",\"$10\",\"$12\",\"$14\",\"$5\",\"$9\",\"$13}';fi;done</proc/diskstats;rm -rf \"$f\""

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_os', methods=['GET'])
    def get_data_ssm_os():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "if [ $(cat /etc/os-release | sed -n 's/^PRETTY_NAME=\"\(.*\)\"/\\1/p') != '' ]; then cat /etc/os-release | sed -n 's/^PRETTY_NAME=\"\(.*\)\"/\\1/p'; else cat /etc/os-release | sed -n 's/^NAME=\"\(.*\)\"/\\1/p'; fi"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_connections', methods=['GET'])
    def get_data_ssm_connections():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "if command -v netstat >/dev/null 2>&1; then netstat -tuan | awk '{$2=$3=\"\"; print $0}' | tail -n +3 | sort | tr -s ' ' ; else echo 'net-tools is not installed'; fi"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_last', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_last():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "last | head -n -2"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_w', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_w():

        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        command = "w -h | tr -s ' '"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_updatespackages', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_updatespackages():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "if command -v apt >/dev/null 2>&1; then echo 'apt' > /tmp/" + formateduuid + "; apt list --upgradable 2>> /dev/null | tail -n +2 | tr -s '\\n' | tr -d '[]' | sed 's|/[^ ]*||' | awk '{printf \" %s,%s,%s,\", $1,$2, $3; for (i=4; i<=NF; ++i) printf \"%s \", $i; printf \"\\n\"}' >> /tmp/" + formateduuid + "&& du -k /tmp/" + formateduuid + " | cut -f1; elif command -v yum >/dev/null 2>&1; then echo 'yum' > /tmp/" + formateduuid + "; script -q -c 'yum check-update' /dev/null | grep -P '.*\.\w*\s.*\s.*' | tr -s ' ' | sed 's/.$//' | awk 'NF == 3' >> /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1; else echo 'dnf' > /tmp/" + formateduuid + "; script -q -c 'dnf check-update' /dev/null | grep -P '.*\.\w*\s.*\s.*' | tr -s ' ' | sed 's/.$//' | awk 'NF == 3' >> /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1; fi"

        file_kb = send_command(instanceid, command, region, jsonifyno = "no")

        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)

    @app.route('/get_data_ssm_installedpackages', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_installedpackages():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "if command -v dpkg >/dev/null 2>&1; then echo 'dpkg' > /tmp/" + formateduuid + "; dpkg -l | tail -n +6 | awk '{printf \"%s,%s,%s,\", $2, $3, $4; for (i=5; i<=NF; ++i) printf \"%s \", $i; printf \"\\n\"}' >> /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1; else echo 'rpm' > /tmp/" + formateduuid + "; rpm -qa --qf '%{NAME},%{VERSION},%{RELEASE},%{ARCH},%{INSTALLTIME:date},%{BUILDTIME:date},%{SIZE},%{LICENSE},%{VENDOR},%{URL},%{SUMMARY},%{PACKAGER}\\n' | sort >> /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1; fi"

        file_kb = send_command(instanceid, command, region, jsonifyno = "no")

        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)

    @app.route('/get_data_ssm_services', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_services():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        command = "if command -v systemctl >/dev/null 2>&1; then systemctl --type=service | sed 's/^[ \\t]*//' | awk '{ print $1 }' | grep '.service' | sed 's/.service//g'; else service --status-all | sed 's|.* ||'; fi"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_services_detailsloop', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_services_detailsloop():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        subfilteredserviceslines = request.args.get("subfilteredserviceslines").split(',')
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "if command -v systemctl >/dev/null 2>&1; then "

        status_lines_systemctl = [f"systemctl status {line} >> /tmp/{formateduuid}; echo 'x&-z$' >> /tmp/{formateduuid}" for line in subfilteredserviceslines]
        command += "; ".join(status_lines_systemctl[:-1]) + "; " + status_lines_systemctl[-1]
        command += "; else "

        status_lines_service = [f"service {line} status >> /tmp/{formateduuid}; echo 'x&-z$' >> /tmp/{formateduuid}" for line in subfilteredserviceslines]
        command += "; ".join(status_lines_service[:-1]) + "; " + status_lines_service[-1]
        command += "; fi && du -k /tmp/" + formateduuid + " | cut -f1"
        
        file_kb = send_command(instanceid, command, region, jsonifyno = "no")

        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)

    @app.route('/get_data_ssm_services_details', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_services_details():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")

        subfilteredserviceslines_str = request.args.get("subfilteredserviceslines")
        subfilteredserviceslines = json.loads(subfilteredserviceslines_str)

        command = "if command -v systemctl >/dev/null 2>&1; then "
        command += " ; echo 'x&-z$'; ".join([ f"systemctl status {line}" for line in subfilteredserviceslines])
        command += "; else "
        command += " ; echo 'x&-z$'; ".join([ f"service {line} status" for line in subfilteredserviceslines])
        command += "; fi;"

        return send_command(instanceid, command, region)

    @app.route('/get_data_ssm_lsof', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_lsof():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "if command -v lsof >/dev/null 2>&1; then for i in $(ps -ax | tail -n +2 | awk '{print $1}'); do lsof -p \"$i\"; echo \"x&-z$\"; done; else echo \"lsof is not installed\"; fi > /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1"

        file_kb = send_command(instanceid, command, region, jsonifyno = "no")

        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)
        
    @app.route('/get_data_ssm_processpermises', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_processpermises():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "ps -eo pid,user,fuser,fgroup,uid,gid,euid,egid,rgid,fgid,fuid,suid,sgid,luid,ouid,pgid,ppid,supgid,supgrp,comm | tail -n +2 | tr -s ' ' > /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1 "

        file_kb = send_command(instanceid, command, region, jsonifyno = "no")
        
        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)

    @app.route('/get_data_ssm_processdata', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_processdata():
        instanceid = request.args.get("instanceid")
        region = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"
        
        command = "ps -eo pid,user,uid,session,tty,start,time,stat,psr,nice,pri,rtprio,pcpu,pmem,vsz,rss,flags,wchan:20,args | tail -n +2 | tr -s ' ' > /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1"

        file_kb = send_command(instanceid, command, region, jsonifyno = "no")

        if len(file_kb) == 0:
            return "None"
        else:
            return loop_command(int(file_kb), instanceid, formateduuid, region)
        
    @app.route('/get_data_ssm_monitor', methods=['GET'])
    @login_required
    @ssm_enabled
    def get_data_ssm_monitor():
        INSTANCEID = request.args.get("instanceid")
        REGION = request.args.get("region")
        formateduuid = str(uuid.uuid4()) + ".txt"

        command = "{ top -bn1 | awk '/Cpu/ { print $2}' | tr -d '\n'; echo -n ';' ; free | grep Mem | awk '{print $3/$2 * 100.0}' | tr -d '\n'; echo -n ';' ; w -i -h | tr -s ' ' | awk '{ print $1, $2, $3, $4}'; echo -n ';' ; journalctl | egrep 'useradd|usermod|userdel|groupadd|groupmod|groupdel' | awk '/useradd.*new user/ {gsub(/name=| UID=| GID=| home=| shell=| from=/,\"\"); print \"useradd,\"$1\" \"$2\" \"$3\",\"$4\",\"$8}; /useradd.*new group/ {gsub(/name=/,\"\"); gsub(/, GID=/,\",\"); print \"groupadd,\"$1\" \"$2\" \"$3\",\"$4\",\"$8,$9}; /usermod.*change user/ {printf \"usermod,\"$1\" \"$2\" \"$3\",\"$4\",\"$8\",\"; for (i=6; i<=NF; i++) printf $i\" \"; print \"\"}; /userdel.*delete user/ {print \"userdel,\"$1\" \"$2\" \"$3\",\"$4\",\"$8}; /groupadd.*new group/ {gsub(/name=| GID=/,\"\"); print \"groupadd,\"$1\" \"$2\" \"$3\",\"$4\",\"$8}; /groupmod.*group changed in \/etc\/group/ {printf \"groupmod,\"$1\" \"$2\" \"$3\",\"$4\",\"; for (i=6; i<=NF; i++) printf $i\" \"; print \"\"}; /groupdel.*removed$/ {printf \"groupdel,\"$1\" \"$2\" \"$3\",\"$4\",\"$7\",\"; for (i=6; i<=NF; i++) printf $i\" \"; print \"\"};' | tr -d \"'\"; echo -n ';' ; netstat -ntu | awk '{$2=$3=\"\"; print $0}' | tail -n +3 | sort | tr -s ' ' | awk '{print $2, $3}' | tr ' ' '\n' | rev | sed 's/:/ /' | rev | xargs -n 4; echo -n ';' ; mount | awk '{ print $1, $3, $5, $6}'; echo -n ';' ; lastb -w | head -n -2 | sed 's/^ /(no-user)/' | awk 'BEGIN{OFS=\",\"} {out=$1 OFS $2 OFS $3 OFS $4; for (i=5; i<=NF; i++) out=out\" \"$i; print out}' ; echo ';'; systemctl --type=service -all | tr -d '●' | sed 's/^[ \\t]*//' | awk '{ print $1, $2, $4 }' | grep '.service' | sed 's/.service//g'; echo ';'; systemctl list-unit-files --type=service | awk '{print $1}' | grep '.service' | sed 's/.service//g'; echo ';'; pidstat -d | awk 'NR>2' | sed 's/^[ \\t]*//' ; echo ';'; } >> /tmp/" + formateduuid + " && du -k /tmp/" + formateduuid + " | cut -f1"

        file_kb = send_command(INSTANCEID, command, REGION, jsonifyno = "no")
        
        if len(file_kb) == 0:
            return "None"
        else:
            session[INSTANCEID + "-" + REGION] = str(uuid.uuid4())

            DATA = loop_command(int(file_kb), INSTANCEID, formateduuid, REGION, jsonifyno = "no")
             
            payload = {
                "timestamp": datetime.datetime.utcnow().isoformat(),
                "expiration": (datetime.datetime.utcnow() + datetime.timedelta(seconds=9)).isoformat(),
                "random": random.random(),
                "instance_id": INSTANCEID,
                "region": REGION,
                "UUID": hashlib.sha256(session[INSTANCEID + "-" + REGION].encode()).hexdigest()
            }

            token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")

            return jsonify({"data": DATA, "token": token})
        
    # Validate token for SSM History
    def validate_token():

        token = request.headers.get('Authorization')
        if not token:
            return False

        token = token.replace('Bearer ', '') 
        
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])

            # Check fields
            REQUIRED_FIELDS = ["timestamp", "expiration", "random", "instance_id", "region", "UUID"]
            if not all(field in payload for field in REQUIRED_FIELDS):
                return False

            # Check times
            current_time = datetime.datetime.utcnow()
            timestamp = datetime.datetime.fromisoformat(payload["timestamp"])
            expiration = datetime.datetime.fromisoformat(payload["expiration"])
            if expiration - timestamp > datetime.timedelta(seconds=10):
                return False
            
            if current_time > expiration or expiration - current_time > datetime.timedelta(seconds=10):
                return False

            # Check random
            if not 0 < payload["random"] < 1:
                return False

            # Check session field
            session_key = f"{payload['instance_id']}-{payload['region']}"
            if session_key not in session:
                return False

            # Check UUID
            if payload["UUID"] != hashlib.sha256(session[session_key].encode()).hexdigest():
                return False

            return True
        
        except jwt.ExpiredSignatureError:
            return False
        
        except jwt.InvalidTokenError:
            return False
    
    # Validate data for SSM History
    def validate_data(data):

        REQUIRED_FIELDS = ['date', 'instanceid', 'event', 'details']
        if not all(field in data for field in REQUIRED_FIELDS):
            return False

        if len(data) != len(REQUIRED_FIELDS):
            return False

        if not all(isinstance(value, str) for value in data.values()):
            return False

        for value in data.values():
            if any(substring in value for substring in ['<script>', 'exec(', 'os.']):
                return False

        date_format = '%Y-%m-%d %H:%M:%S:%f'
        data_date = datetime.datetime.strptime(data['date'], date_format)
        current_time = datetime.datetime.utcnow()
        time_difference = current_time - data_date

        if time_difference.total_seconds() > 15:
            return False

        return True
    
    @app.route('/send_data_ssm_history', methods=['POST'])
    @login_required
    @ssm_enabled
    def send_data_ssm_history():

        if not 'uuid' in session and not session.get('uuid'):
            path = request.path
            return render_template('login/404.html', path=path), 404
        
        token = request.headers.get('Authorization')

        if token is None:
            return jsonify({"error": "PAGE NOT FOUND"}), 404

        if validate_token():
            try:
                data = request.json.get('data')
                if validate_data(data):
                    DATE = datetime.datetime.strptime(data['date'], '%Y-%m-%d %H:%M:%S:%f')
                    query = History(id=session['uuid'], date=DATE, instance_id=data['instanceid'], event=data['event'], details=data['details'])
                    db.session.add(query)
                    db.session.commit()
                    return jsonify({"message": "OK"}), 200
                else:
                    return jsonify({"error": "INVALID DATA"}), 400
            except Exception as e:
                return jsonify({"error": "INVALID DATA"}), 400
        else:
            return jsonify({"error": "UNAUTHORIZED"}), 401
        
    @app.route('/export_data_ssm_history_excel', methods=['GET'])
    @login_required
    @ssm_enabled
    def export_data_smm_history_excel():

        if not 'uuid' in session and not session.get('uuid'):
            path = request.path
            return render_template('login/404.html', path=path), 404
        
        DATABASEDATA = db.session.query(History).filter_by(id=session['uuid']).all()
        
        # Create Excel and select sheet
        wb = Workbook()
        ws = wb.active

        # Headers
        ws.append(['Date', 'Instance ID', 'Event', 'Details'])
        header_style = NamedStyle(name='Header Style', font=Font(size=12, bold=True))
        for cell in ws['1:1']:
            cell.style = header_style

        # Data
        for each in DATABASEDATA:
            row_data = [each.date.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3], each.instance_id, each.event, each.details]
            ws.append(row_data)

        for cell in ws['A'][1:]:
            cell.number_format = 'yyyy-mm-dd h:mm:ss.000'

        # Create table
        tab = Table(displayName="Table1", ref="A1:D{}".format(len(DATABASEDATA) + 1))

        # Add table style
        style = TableStyleInfo(
            name="TableStyleLight13", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # Add table to style
        ws.add_table(tab)

        # Auto adjust width column
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save Excel book
        excel_output = BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        return send_file(
            excel_output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name='history-' + (lambda date: date.strftime("%Y-%m-%d %H:%M:%S"))(datetime.datetime.now()) + '.xlsx'
        )
    
    @app.route('/export_data_ssm_history_csv', methods=['GET'])
    @login_required
    @ssm_enabled
    def export_data_smm_history_csv():

        if not 'uuid' in session and not session.get('uuid'):
            path = request.path
            return render_template('login/404.html', path=path), 404

        DATABASEDATA = db.session.query(History).filter_by(id=session['uuid']).all()
        
        # Create in-memory file
        csv_data = io.StringIO()

        # Create CSV writer
        csv_writer = csv.writer(csv_data)

        # Write header
        csv_writer.writerow(['date', 'instance_id', 'event', 'details'])

        # Write data
        for each in DATABASEDATA:
            csv_writer.writerow([
                each.date.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3],
                str(each.instance_id),
                each.event,
                each.details
            ])

        # Move to the beginning of the in-memory file
        csv_data.seek(0)

        # Encode the entire content to bytes
        csv_binary_data = io.BytesIO(csv_data.getvalue().encode('utf-8'))
        
        return send_file(
            csv_binary_data,
            mimetype='text/csv',
            as_attachment=True,
            download_name='history-' + (lambda date: date.strftime("%Y-%m-%d %H:%M:%S"))(datetime.datetime.now()) + '.csv'
        )
        